import os
import requests
import logging
import pandas as pd
from bs4 import BeautifulSoup as Bs, BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from pathlib import Path
from datetime import datetime
import pytz
import sys
import smtplib
from email.message import EmailMessage
import gzip
import shutil
import time
from requests.exceptions import RequestException


# Activate '.env' file
env_path = Path('.') / '.env'
load_dotenv(dotenv_path=env_path)

# Logging setup
log_file = "nfl_spread_script.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)

# Dry-run toggle
DRY_RUN = False

# NFL team abbreviations
team_abbr = {
    "49ERS": "SF", "BEARS": "CHI", "BENGALS": "CIN", "BILLS": "BUF",
    "BRONCOS": "DEN", "BROWNS": "CLE", "BUCCANEERS": "TB", "CARDINALS": "ARI",
    "CHARGERS": "LAC", "CHIEFS": "KC", "COLTS": "IND", "COMMANDERS": "WAS",
    "COWBOYS": "DAL", "DOLPHINS": "MIA", "EAGLES": "PHI", "FALCONS": "ATL",
    "GIANTS": "NYG", "JAGUARS": "JAX", "JETS": "NYJ", "LIONS": "DET",
    "PACKERS": "GB", "PANTHERS": "CAR", "PATRIOTS": "NE", "RAIDERS": "LV",
    "RAMS": "LAR", "RAVENS": "BAL", "SAINTS": "NO", "SEAHAWKS": "SEA",
    "STEELERS": "PIT", "TEXANS": "HOU", "TITANS": "TEN", "VIKINGS": "MIN"
}


def send_error_email(subject, body, log_path):
    try:
        msg = EmailMessage()
        msg["From"] = os.getenv("EMAIL_ADDRESS")
        msg["To"] = os.getenv("TO_EMAIL_ADDRESS")
        msg["Subject"] = subject
        msg.set_content(body)

        # Attach log file
        with open(log_path, "rb") as f:
            msg.add_attachment(f.read(), maintype="text", subtype="plain", filename=os.path.basename(log_path))

        with smtplib.SMTP(os.getenv("SMTP_SERVER"), int(os.getenv("SMTP_PORT"))) as server:
            server.starttls()
            server.login(os.getenv("EMAIL_ADDRESS"), os.getenv("EMAIL_PASSWORD"))
            server.send_message(msg)

        logging.info("Error email sent successfully.")
    except Exception as e:
        logging.warning(f"Failed to send error email: {e}")

def send_test_email():
    subject = "NFL Automation Test Email"
    body = "This is a test email to confirm Gmail alert functionality is working."
    try:
        send_error_email(subject, body, log_file)
        logging.info("Test email sent successfully.")
    except Exception as e:
        logging.critical(f"Test email failed: {e}", exc_info=True)


def archive_log_file():
    try:
        log_path = log_file  # e.g., "nfl_spread_script.log"
        timestamp = datetime.now().strftime("%Y-%m-%d")
        archive_name = f"logs/nfl_spread_script_{timestamp}.log.gz"

        os.makedirs("logs", exist_ok=True)

        with open(log_path, "rb") as f_in:
            with gzip.open(archive_name, "wb") as f_out:
                shutil.copyfileobj(f_in, f_out)

        logging.info(f"Archived log to {archive_name}")

        # Optional: clear original log file
        open(log_path, "w").close()
        logging.info("Cleared original log file after archiving.")

    except Exception as e:
        logging.error(f"Failed to archive log file: {e}")
        send_error_email(
            subject="NFL Spread Script: ERROR - Log Archiving Failed",
            body=f"Failed to archive log file:\n{e}",
            log_path=log_path
        )


def fetch_with_retry(url, headers=None, max_retries=3, backoff_factor=2, timeout=10):
    attempt = 0
    while attempt < max_retries:
        try:
            response = requests.get(url, headers=headers, timeout=timeout)
            response.raise_for_status()
            return response
        except RequestException as e:
            attempt += 1
            wait_time = backoff_factor ** attempt
            logging.warning(f"Request failed (attempt {attempt}/{max_retries}): {e}. Retrying in {wait_time}s...")
            time.sleep(wait_time)
    logging.error(f"All {max_retries} attempts failed for URL: {url}")
    raise ConnectionError(f"Failed to fetch data from {url} after {max_retries} retries.")


def get_webpage(url, headers=None):
    try:
        response = fetch_with_retry(url, headers=headers)
        return BeautifulSoup(response.content, "html.parser")
    except Exception as e:
        logging.error(f"Failed to fetch webpage after retries: {e}")
        return None


def get_week_number(soup):
    try:
        return soup.find("div", class_="filters-week-picker") \
            .find("div", class_="selector week-picker-week") \
            .find("li", class_="menu-item active") \
            .find("span", attrs={"data-endpoint": True}).get_text()
    except AttributeError:
        msg = "Week number not found in HTML structure."
        logging.warning(msg)
        send_error_email(
            subject="NFL Spread Script: ERROR - Week Number Missing",
            body=msg,
            log_path=log_file
        )
        return "Unknown"


def extract_team_info(table, side):
    tr = table.find("tr", attrs={"data-side": side})
    name = tr.find("span", class_="team-name").find("a").find("span").get_text().upper()
    abbr = tr.find("span", class_="team-name").find("a", attrs={"data-abbr": True}).get("data-abbr")
    return name, abbr


def extract_spread_and_favorite(table):
    td = table.find("td", attrs={"data-field": "current-spread"})
    if not td:
        return "TBD", None
    span = td.find("span", class_="data-value")
    raw = span.get_text(strip=True) if span else td.get_text(strip=True).split(" ")[0]
    if raw.lower() in ["tbd", "n/a", ""]:
        return "TBD", None
    raw_clean = raw.replace("−", "-").replace("+", "").strip()
    side = td.get("data-side")
    return raw_clean, side


def extract_datetime(table):
    # Try real HTML format first
    span = table.find("span", attrs={"data-value": True})
    if span:
        try:
            return datetime.fromisoformat(span.get("data-value"))
        except Exception:
            pass

    # Fallback to mock HTML format
    try:
        date_str = table.find("div", class_="game-date").get_text(strip=True)
        return datetime.strptime(date_str, "%A, %B %d, %Y")
    except Exception as e:
        logging.warning(f"Date parsing failed: {e}")
        return None



def parse_game_card(table):
    away_name, away_abbr = extract_team_info(table, "away")
    home_name, home_abbr = extract_team_info(table, "home")
    spread, favorite_side = extract_spread_and_favorite(table)
    date_time = extract_datetime(table)
    return [away_name, spread, home_name, away_abbr, home_abbr, home_name.upper(), date_time, favorite_side]


def scrape_nfl_data():
    url = "https://www.scoresandodds.com/nfl"
    soup = get_webpage(url)

    if not soup:
        logging.error("Failed to load NFL page.")
        send_error_email(
            subject="NFL Scraper Error: Page Load Failure",
            body="Failed to load NFL page from scoresandodds.com.",
            log_path=log_file
        )
        return None, "Unknown"

    try:
        week = get_week_number(soup)
        logging.info(f"Scraping data for Week {week}")
    except Exception as e:
        logging.error(f"Failed to extract week number: {e}", exc_info=True)
        send_error_email(
            subject="NFL Scraper Error: Week Extraction Failed",
            body=f"Error extracting week number:\n{e}",
            log_path=log_file
        )
        return None, "Unknown"

    data = []
    finalized_count = 0
    pending_count = 0

    for table in soup.find_all("div", class_="event-card"):
        try:
            row = parse_game_card(table)
            if row[1] == "TBD":
                pending_count += 1
            else:
                finalized_count += 1
            data.append(row)
        except Exception as e:
            logging.warning(f"Failed to parse game card: {e}")

    if not data:
        logging.error("No game data found.")
        send_error_email(
            subject="NFL Scraper Error: No Game Data",
            body="Scraper ran successfully but found no game data.",
            log_path=log_file
        )
        return None, week

    try:
        df = pd.DataFrame(data, columns=[
            "Team1", "Spread", "Team2", "Team1_Abbr", "Team2_Abbr",
            "Home_Team", "UTC_DateTime", "Favorite_Side"
        ])
        df = apply_team_abbreviations(df)
        logging.info(f"Scraped {len(df)} games: {finalized_count} finalized, {pending_count} pending")
        return df, week
    except Exception as e:
        logging.critical(f"DataFrame construction or abbreviation failed: {e}", exc_info=True)
        send_error_email(
            subject="NFL Scraper Critical Error: DataFrame Failure",
            body=f"Critical failure during DataFrame construction or abbreviation:\n{e}",
            log_path=log_file
        )
        return None, week


def apply_team_abbreviations(df):
    df["Team1"] = df["Team1"].str.upper()
    df["Team2"] = df["Team2"].str.upper()
    df["Team1_Abbr"] = df["Team1"].map(team_abbr)
    df["Team2_Abbr"] = df["Team2"].map(team_abbr)

    missing_team1 = df[df["Team1_Abbr"].isna()]["Team1"].unique()
    missing_team2 = df[df["Team2_Abbr"].isna()]["Team2"].unique()
    missing = list(missing_team1) + list(missing_team2)

    if missing:
        logging.warning(f"Missing abbreviations for: {missing}")
        if len(missing) > 3:
            send_error_email(
                subject="NFL Spread Script: ERROR - Abbreviation Mapping",
                body=f"Missing team abbreviations for: {missing}",
                log_path=log_file
            )
    return df


def extract_favorite_underdog(row):
    spread_val = row["Spread"]
    favorite_side = row["Favorite_Side"]
    team1 = row["Team1"]
    team2 = row["Team2"]
    abbr1 = row["Team1_Abbr"]
    abbr2 = row["Team2_Abbr"]

    if spread_val == "TBD" or favorite_side not in ["home", "away"]:
        return "TBD", "TBD", 0.0, None, None

    try:
        spread_float = float(spread_val)
    except ValueError:
        return "TBD", "TBD", 0.0, None, None

    if favorite_side == "home":
        spread_team = team2
        spread_abbr = abbr2
        other_team = team1
        other_abbr = abbr1
    else:
        spread_team = team1
        spread_abbr = abbr1
        other_team = team2
        other_abbr = abbr2

    if spread_float < 0:
        favorite = spread_team
        underdog = other_team
        fav_abbr = spread_abbr
        und_abbr = other_abbr
        spread_display = abs(spread_float)
    else:
        favorite = other_team
        underdog = spread_team
        fav_abbr = other_abbr
        und_abbr = spread_abbr
        spread_display = spread_float

    return favorite, underdog, spread_display, fav_abbr, und_abbr


def filter_games_by_day(df):
    pacific = pytz.timezone("America/Los_Angeles")
    now = datetime.now(pacific)
    dotw = now.strftime("%A")

    # ✅ Ensure UTC_DateTime is a datetime object
    df["UTC_DateTime"] = pd.to_datetime(df["UTC_DateTime"], errors="coerce")

    # ✅ Filter out games that have already started
    df_filtered = df[df["UTC_DateTime"] > now].copy()
    excluded = df[df["UTC_DateTime"] <= now]

    logging.info(f"Excluded {len(excluded)} played games for {dotw}:")
    for _, row in excluded.iterrows():
        game_day = (
            row["Local_DateTime"].strftime("%A")
            if pd.notna(row.get("Local_DateTime"))
            else "Unknown Day"
        )
        logging.info(f"  {row['Team1']} vs {row['Team2']} on {game_day} ({row['UTC_DateTime']})")

    if not df_filtered.empty and "Local_DateTime" in df_filtered.columns:
        earliest_game = df_filtered["Local_DateTime"].min()
        logging.info(f"Earliest remaining game is on {earliest_game.strftime('%A, %Y-%m-%d %I:%M %p')}")
    logging.info(f"Filtered games for {dotw}: {len(df_filtered)} remaining")

    return df_filtered, dotw

def get_local_day(utc_str):
    pacific = pytz.timezone("America/Los_Angeles")
    try:
        dt = datetime.strptime(utc_str, "%Y-%m-%dT%H:%M:%SZ")
        local_dt = dt.replace(tzinfo=pytz.utc).astimezone(pacific)
        return local_dt.strftime("%A")
    except Exception as e:
        logging.warning(f"Failed to parse UTC datetime: {e}")
        send_error_email(
            subject="NFL Spread Script: ERROR - UTC Day Conversion",
            body=f"Failed to convert UTC string: {utc_str}\nError: {e}",
            log_path=log_file
        )
        return "Unknown"



def get_local_datetime(utc_str):
    pacific = pytz.timezone("America/Los_Angeles")
    try:
        dt = datetime.strptime(utc_str, "%Y-%m-%dT%H:%M:%SZ")
        return dt.replace(tzinfo=pytz.utc).astimezone(pacific)
    except Exception as e:
        logging.warning(f"Failed to convert UTC datetime: {e}")
        send_error_email(
            subject="NFL Spread Script: ERROR - UTC Datetime Conversion",
            body=f"Failed to convert UTC string: {utc_str}\nError: {e}",
            log_path=log_file
        )
        return None


def update_excel(wk_number, df_filtered, dotw):
    try:
        file = os.getenv("file_path")
        wb = load_workbook(filename=file)
        all_sheets = wb.sheetnames
        template = wb.worksheets[0]

        home_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
        clear_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Create or overwrite sheet
        if wk_number in all_sheets:
            new_wk_sheet = wb[wk_number]
            logging.info(f"Overwriting existing sheet: {wk_number}")
        else:
            template_copy = wb.copy_worksheet(template)
            template_copy.title = wk_number
            new_wk_sheet = wb[wk_number]
            logging.info(f"Created new sheet: {wk_number}")

        # Activate the new sheet
        for sheet in wb:
            sheet.views.sheetView[0].tabSelected = False
        wb.active = new_wk_sheet
        new_wk_sheet.views.sheetView[0].tabSelected = True

        # Defensive check for Excel_Row
        if "Excel_Row" not in df_filtered.columns:
            msg = "Excel_Row column missing from DataFrame. Aborting Excel update."
            logging.critical(msg)
            send_error_email(
                subject="NFL Excel Update Critical Error",
                body=msg,
                log_path=log_file
            )
            return

        df = df_filtered.copy()
        df = df[df["Excel_Row"].notna()]
        df["Excel_Row"] = df["Excel_Row"].astype(int)

        rows_to_update = df["Excel_Row"].unique()

        # Clear all rows that will be updated
        for row in rows_to_update:
            logging.info(f"Clearing row {row}")
            for col in [3, 4, 5, 9, 11, 14, 15]:
                new_wk_sheet.cell(row=row, column=col).value = None
                new_wk_sheet.cell(row=row, column=col).fill = clear_fill

        # Update each row with FAVORITE vs UNDERDOG
        for _, row in df.iterrows():
            try:
                excel_row = int(row["Excel_Row"])
                favorite, underdog, spread_val, fav_abbrev, und_abbr = extract_favorite_underdog(row)
                ht = row["Home_Team"]

                logging.info(f"Updating row {excel_row}: {favorite} vs {underdog}, spread {spread_val}")

                new_wk_sheet.cell(row=excel_row, column=3).value = favorite
                new_wk_sheet.cell(row=excel_row, column=4).value = spread_val
                new_wk_sheet.cell(row=excel_row, column=5).value = underdog
                new_wk_sheet.cell(row=excel_row, column=9).value = fav_abbrev
                new_wk_sheet.cell(row=excel_row, column=11).value = und_abbr

                cell_c = new_wk_sheet.cell(row=excel_row, column=3)
                cell_e = new_wk_sheet.cell(row=excel_row, column=5)

                if favorite == ht:
                    cell_c.fill = home_fill
                    cell_e.fill = clear_fill
                elif underdog == ht:
                    cell_e.fill = home_fill
                    cell_c.fill = clear_fill
                else:
                    cell_c.fill = clear_fill
                    cell_e.fill = clear_fill
                    logging.warning(f"Home team '{ht}' not matched in favorite/underdog for row {excel_row}")

                for col in [14, 15]:
                    new_wk_sheet.cell(row=excel_row, column=col).fill = clear_fill

            except Exception as e:
                logging.warning(f"Error updating row {row.get('Excel_Row', 'Unknown')}: {e}")

        wb.save(file)
        logging.info(f"Excel updated and saved for {wk_number}")

    except Exception as e:
        logging.critical(f"Excel update failed: {e}", exc_info=True)
        try:
            logging.error(f"Available sheets: {wb.sheetnames}")
        except:
            logging.error("Workbook not loaded—no sheet names available.")

        send_error_email(
            subject="NFL Excel Update Critical Error",
            body=f"Excel update failed:\n{e}",
            log_path=log_file
        )

def verify_matchkey_alignment(df_full, df_filtered):
    full_keys = df_full["MatchKey"].drop_duplicates()
    filtered_keys = df_filtered["MatchKey"].drop_duplicates()
    unmatched = filtered_keys[~filtered_keys.isin(full_keys)]

    if unmatched.empty:
        logging.info("All MatchKeys in df_filtered matched df_full.")
    else:
        logging.warning("Unmatched MatchKeys in df_filtered:")
        for key in unmatched:
            logging.warning(f"  {key}")
        send_error_email(
            subject="NFL Spread Script: ERROR - MatchKey Mismatch",
            body=f"Unmatched MatchKeys found:\n" + "\n".join(unmatched),
            log_path=log_file
        )

def normalize_matchkeys(df):
    df["Team1"] = df["Team1"].astype(str).str.strip().str.upper()
    df["Team2"] = df["Team2"].astype(str).str.strip().str.upper()
    df["MatchKey"] = (df["Team1"] + " vs " + df["Team2"]).str.strip().str.upper()
    return df

def assign_excel_rows(df):
    """
    Assigns Excel row numbers based on game weekday.
    Skips Friday games. Thursday starts at row 1.
    """
    weekday_order = ["Thursday", "Saturday", "Sunday", "Monday", "Tuesday", "Wednesday"]
    row_counter = 1
    excel_rows = []

    for dt in df["UTC_DateTime"]:
        weekday = dt.strftime("%A")
        if weekday == "Friday":
            excel_rows.append(None)  # Skip Friday games
        else:
            excel_rows.append(row_counter)
            row_counter += 1

    return excel_rows

def main():
    logging.info("Starting NFL pool automation...")

    try:
        # ✅ Scrape and normalize
        df_raw, week_label = scrape_nfl_data()

        if df_raw is None or not isinstance(df_raw, pd.DataFrame):
            msg = "Scraping failed or returned invalid data. Aborting pipeline."
            logging.critical(msg)
            send_error_email(
                subject="NFL Automation Critical Error: Scraping Failed",
                body=msg,
                log_path=log_file
            )
            return

        logging.info(f"Scraping data for Week {week_label}")
        logging.info(f"Scraped {len(df_raw)} games")

        df_raw = normalize_matchkeys(df_raw)

        # ✅ Count how many games have already started
        now = datetime.now(pytz.timezone("America/Los_Angeles"))
        df_raw["UTC_DateTime"] = pd.to_datetime(df_raw["UTC_DateTime"], errors="coerce")
        excluded_count = len(df_raw[df_raw["UTC_DateTime"] <= now])
        logging.info(f"Detected {excluded_count} played games before {now.strftime('%A %I:%M %p')}")

        # ✅ Assign Excel_Row based on full schedule, offset by excluded games
        df_raw = df_raw.reset_index(drop=True)
        df_raw["Excel_Row"] = df_raw.index + 2 + excluded_count  # Dynamic offset

        # ✅ Filter out played games — Excel_Row is preserved
        df_filtered, dotw = filter_games_by_day(df_raw)

        # ✅ Confirm Excel_Row exists
        if "Excel_Row" not in df_filtered.columns:
            msg = "Excel_Row missing from filtered DataFrame. Aborting."
            logging.critical(msg)
            send_error_email(
                subject="NFL Automation Critical Error: Excel_Row Missing",
                body=msg,
                log_path=log_file
            )
            return

        # ✅ Confirm all rows have Excel_Row
        unmatched = df_filtered[df_filtered["Excel_Row"].isna()]
        if not unmatched.empty:
            logging.warning(f"Unmatched rows after filtering: {len(unmatched)}")
            for _, row in unmatched.iterrows():
                logging.warning(f"  {row['Team1']} vs {row['Team2']} — MatchKey: {row['MatchKey']}")
        else:
            logging.info("[OK] All filtered games have Excel_Row assigned.")

        # ✅ Preview post-filter
        logging.info("Post-filter preview:")
        preview_cols = ["Team1", "Team2", "MatchKey", "Excel_Row"]
        logging.info(df_filtered[preview_cols].to_string(index=False))

        # ✅ Update Excel
        update_excel(week_label, df_filtered, dotw)

        logging.info("NFL pool automation complete.")

    except Exception as e:
        logging.critical(f"Unhandled exception in main(): {e}", exc_info=True)
        send_error_email(
            subject="NFL Automation Crash",
            body=f"Unhandled exception in main():\n{e}",
            log_path=log_file
        )

if __name__ == "__main__":
    main()